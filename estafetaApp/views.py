from cmath import phase
from itertools import count
from django.shortcuts import render, redirect, get_object_or_404
from .forms import *
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse, Http404, HttpResponseNotFound
from django.contrib.auth import authenticate, login, logout
from account.models import Account
from django.core.files.storage import FileSystemStorage
import openpyxl
from .models import Tests, Team
from django.db.models import Q
import xlrd 
import csv
import pandas as pd

def table_page(request):
    content = {}
    if request.user.is_authenticated and request.user.is_admin:
        if request.FILES:
            excel_file = request.FILES["excel_file"]

            wb = openpyxl.load_workbook(excel_file)

            worksheet = wb["Sheet1"]
            print(worksheet)

            excel_data = list()

            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)
            print(excel_data[1::])

            content["excel_data_start"] = excel_data[0]

            content["excel_data"] = excel_data[1::]

            return render(request, 'tables.html', content)
        else:
            return render(request, 'tables.html', content)
    else:
        return HttpResponseRedirect('/')

def profile_page(request):
    content = {}
    if request.user.is_authenticated:
        email = request.user
        person = Account.objects.get(email=email)
        content['user'] = person
        if request.method == 'POST' and 'createTeam' in request.POST:
            try:
                team = Team()
                team.name = request.POST['name']
                team.language = request.POST['language']
                team.password = request.POST['password']
                team.save()
                person.team = request.POST['name']
                person.save()
                return HttpResponseRedirect('/profile/')
            except:
                print("error")
        if request.method == 'POST' and 'leaveTeam' in request.POST:
            person.team = ""
            person.save()
            return HttpResponseRedirect('/profile/')
        if request.method == 'POST' and 'joinTeam' in request.POST:
            team = Team.objects.get(name=request.POST['joinTeam'])
            if team.password == request.POST['password']:
                person.team = request.POST['joinTeam']
                person.save()
            else:
                print("error")
            return HttpResponseRedirect('/profile/')    
        return render(request, 'profile.html', content)
    else:
        return render(request, 'notadmin.html')


def profileTemplate_page(request, name):
    content = {}
    if request.user.is_authenticated:
        path = f"profile/template{name}.html"
        user = Account.objects.get(email=request.user)
        if user.team:
            team = Team.objects.get(name=user.team)
            content['team'] = team
        else:
            content['team'] = ""
        return render(request, path, content)
    else:
        return HttpResponseRedirect('/singin')    


def searchTeam_page(request, name):
    content = {}
    content['teams'] =  Team.objects.filter(name__icontains=name)
    return render(request, 'outpLists/teamslist.html', content)

def searchTest_page(request, name):
    content = {}
    content['tests'] =  Tests.objects.filter(name__icontains=name)
    print(content['tests'])
    return render(request, 'outpLists/testslist.html', content)

def createtest_page(request):
    if request.user.is_authenticated and request.user.is_admin:
        if request.method == 'POST' and 'submitBtn' in request.POST:

            test = Tests()
            test.name = request.POST['name']
            test.subject = request.POST['subject']
            test.level = request.POST['level']
            test.link = request.POST['link']
            test.date_start = request.POST['date_start']
            test.time_start = request.POST['time_start']
            test.date_end = request.POST['date_end']
            test.time_end = request.POST['time_end']
            test.save()
            return HttpResponseRedirect('/createtest/')
        return render(request, 'createtest.html')
    return render(request, 'notadmin.html')

def finishtest_page(request):
    if request.user.is_authenticated and request.user.is_admin:
        if request.method == 'POST':
            print(request.POST)

            test = Tests.objects.get(id = request.POST['testId'])
            test.is_active = False

            print(request.FILES)
            

            if request.FILES:
                print("---------------OK---------------")
                excel_file = request.FILES["excel_file"]

                wb = openpyxl.load_workbook(excel_file)

                worksheet = wb["Sheet1"]
                # print(worksheet)

                excel_data = list()

                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                    excel_data.append(row_data)

                # file_name = FileSystemStorage.save(file.name, file)
                print(excel_data)

               

                with open(str(test.id)+'.csv', 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerows(excel_data)

                url = str(test.id) + '.csv'
                test.results_link = url
                print(url)

            test.save()

            return HttpResponseRedirect('/finishtest/')
        return render(request, 'finishtest.html')
    return render(request, 'notadmin.html')

def index_page(request):
    return render(request, 'index.html')

def main_page(request):
    if (request.user.is_authenticated):
        content = {}
        tests = Tests.objects.all()
        content['tests'] = tests
        return render(request, 'main.html', content)
    else:
        return HttpResponseRedirect('/hello/')

def resultsall_page(request):
    if (request.user.is_authenticated):
        content = {}
        tests = Tests.objects.all()
        content['tests'] = tests
        return render(request, 'results.html', content)
    else:
        return HttpResponseRedirect('/hello/')

def resultstest_page(request, id):
    content = {}
    test = Tests.objects.get(id=id)
    content['test'] = test
 
    try:
        with open(str(test.id)+'.csv', newline='') as f:
            reader = csv.reader(f)
            data = list(reader)


        content["excel_data_start"] = data[0]

        content["excel_data"] = data[1::]

        return render(request, 'resultTestPage.html', content)
    except:
        return HttpResponseRedirect('/')




def info_page(request):
    content = {}
    
    try:
        test = Tests.objects.get()
        user = Account.objects.get(email=request.user)
        content['user'] = user

        if request.FILES:

            # SAVE FILE TO CSV


            excel_file = request.FILES["excel_file"]

            wb = openpyxl.load_workbook(excel_file)

            worksheet = wb["Sheet1"]
            # print(worksheet)

            excel_data = list()

            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)

            url = str(test.id) + '.csv'
            print(url)

            with open(url, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerows(excel_data)

            test['results_link'] = url

            print(test.id)


            # OPEN CSV FILE 


            with open(test['results_link'], newline='') as f:
                reader = csv.reader(f)
                data = list(reader)
            print(data)


            content["excel_data_start"] = data[0]

            content["excel_data"] = data[1::]

        return render(request, 'info.html', content)
    except:
        return render(request, 'info.html', content)

# def main_page(request):
#     content = {}
#     tests = Tests.objects.all()
#     content['tests'] = tests
#     return render(request, 'main.html', content)

def logout_view(request):
    logout(request)
    return HttpResponseRedirect("/")

def reg_page(request):
    form = SignUpForm(request.POST)
    context = {
        'form': form
    }
    # вход

    if request.method == 'POST' and 'profile_saver3' in request.POST:
        # print(form.errors)
        # print(form.is_valid())
        list = request.POST.getlist('sports')

        print(list)


        if form.is_valid():
            form.save()
            email = form.cleaned_data.get('email')
            first_name = form.cleaned_data.get('first_name')
            password = form.cleaned_data.get('password1')

            phone = request.POST.get('phone')
            country = request.POST.get('country')
            subjects = request.POST.getlist('sports')
            knowledge = request.POST.get('knowledge')

            user = authenticate(email=email, password=password,
             first_name=first_name)

            user.phone = phone
            user.country = country
            user.knowledge = knowledge
            user.subjects = subjects
            print(request.POST)
            user.save()
            return redirect('/login/')

  

    return render(request, 'signin.html', context)

def login_page(request):

    if request.method == 'POST' and 'btnform2' in request.POST:
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request, email=email, password=password)

        if user is not None:
            login(request, user)
            return redirect('/')
        else:
            print('Try again! username or password is incorrect')
    return render(request, 'login.html')